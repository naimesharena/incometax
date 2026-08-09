"""Capital gains engine for FY 2025-26 (AY 2026-27).

Handles every class an individual salaried investor needs:
  * Listed equity / equity-MF with STT: STCG u/s 111A (20%) and LTCG u/s 112A
    (12.5% above the Rs. 1,25,000 annual exemption) incl. the 31-01-2018
    grandfathering rule,
  * Land / building: 24-month holding, resident option of 12.5% without
    indexation vs 20% with indexation (property acquired before 23-07-2024),
  * Other assets STCG (slab rate) and LTCG (12.5% w/o indexation after
    23-07-2024, else 20% indexed),
  * Exemptions u/s 54 / 54EC / 54EE / 54F (with caps),
  * Intra-head loss set-off (ST loss → ST/LT gains; LT loss → LT gains only),
  * Quarter-wise accrual split for interest u/s 234C,
  * Feeds the ITR-1 aggregate (112A within the exemption limit) or the full
    ITR-2 schedules.
"""

from __future__ import annotations

import csv
import io
import json

from .errors import field_scope, parse_float, parse_int

# Cost inflation index (base 2001-02 = 100)
CII = {
    2001: 100, 2002: 105, 2003: 109, 2004: 113, 2005: 117, 2006: 122,
    2007: 129, 2008: 137, 2009: 148, 2010: 167, 2011: 184, 2012: 200,
    2013: 220, 2014: 240, 2015: 254, 2016: 264, 2017: 272, 2018: 280,
    2019: 289, 2020: 301, 2021: 317, 2022: 331, 2023: 348, 2024: 363,
    2025: 376,
}
CII_SALE_YEAR = CII[2025]
GRANDFATHER_DATE = "2018-01-31"
RATE_CHANGE_DATE = "2024-07-23"
EXEMPT_112A = 125000
RATE_111A = 0.20
RATE_112A = 0.125
RATE_LTCG_OTHER = 0.125
RATE_LTCG_INDEXED = 0.20
CAP_54EC = 5000000
CAP_54EE = 5000000

QUARTERS = [("Q1", "2025-04-01", "2025-06-15"), ("Q2", "2025-06-16", "2025-09-15"),
            ("Q3", "2025-09-16", "2025-12-15"), ("Q4", "2025-12-16", "2026-03-15"),
            ("Q5", "2026-03-16", "2026-03-31")]


def _g(m, key, default=0, label=None):
    if m is None or key not in m:
        return default
    return parse_int(m.get(key), default=default, field=label, key_path=str(key))


def _f(m, key, default=0.0, label=None):
    if m is None or key not in m:
        return default
    return parse_float(m.get(key), default=default, field=label, key_path=str(key))


def _valid_date(s):
    s = str(s or "").strip()
    return s if len(s) == 10 and s[4] == "-" and s[7] == "-" else ""


def _quarter(sale_date: str) -> str:
    for q, lo, hi in QUARTERS:
        if sale_date and lo <= sale_date <= hi:
            return q
    return "Q5"


def _holding_months(buy: str, sale: str) -> float:
    if not (buy and sale):
        return 0.0
    try:
        import datetime as dt
        return (dt.date.fromisoformat(sale) - dt.date.fromisoformat(buy)).days / 30.4375
    except ValueError:
        return 0.0


def _indexed_cost(cost: float, buy_year: int) -> float:
    base = CII.get(buy_year)
    if not base or buy_year >= 2025:
        return cost
    return cost * CII_SALE_YEAR / base


# ---------------------------------------------------------------------------
# Row-level computations
# ---------------------------------------------------------------------------

def _equity_lt_row(r: dict, idx: int) -> dict:
    """LTCG u/s 112A per scrip with grandfathering."""
    with field_scope(f"Equity LTCG row {idx} ({r.get('scrip') or 'scrip'})"):
        sale = _g(r, "sale_consideration")
        cost = _g(r, "cost_of_acquisition")
        fmv = _g(r, "fmv_31_01_2018")
        exp = _g(r, "expenses")
    buy = _valid_date(r.get("buy_date"))
    sold = _valid_date(r.get("sale_date"))
    coa = cost
    grand = ""
    if buy and buy <= GRANDFATHER_DATE and fmv > 0:
        coa = max(cost, min(fmv, sale))
        grand = (f"Grandfathered COA = higher of actual Rs. {cost:,} and lower of "
                 f"FMV Rs. {fmv:,} / sale Rs. {sale:,} → Rs. {coa:,}")
    gain = sale - coa - exp
    return {"scrip": str(r.get("scrip") or f"Scrip {idx}"), "isin": str(r.get("isin") or ""),
            "buy_date": buy, "sale_date": sold, "sale": sale, "cost": cost,
            "fmv": fmv, "coa": coa, "expenses": exp, "gain": gain,
            "note": grand, "quarter": _quarter(sold)}


def _equity_st_row(r: dict, idx: int) -> dict:
    with field_scope(f"Equity STCG row {idx} ({r.get('scrip') or 'scrip'})"):
        sale = _g(r, "sale_consideration")
        cost = _g(r, "cost_of_acquisition")
        exp = _g(r, "expenses")
    sold = _valid_date(r.get("sale_date"))
    return {"scrip": str(r.get("scrip") or f"Scrip {idx}"), "buy_date": _valid_date(r.get("buy_date")),
            "sale_date": sold, "sale": sale, "cost": cost, "expenses": exp,
            "gain": sale - cost - exp, "quarter": _quarter(sold)}


def _land_row(r: dict, idx: int) -> dict:
    """Land/building: LTCG if held ≥24m; resident can pick indexation option."""
    with field_scope(f"Land/building row {idx} ({r.get('description') or 'property'})"):
        sale = _g(r, "sale_consideration")
        cost = _g(r, "cost_of_acquisition")
        improv = _g(r, "cost_of_improvement")
        exp = _g(r, "expenses")
    buy = _valid_date(r.get("buy_date"))
    sold = _valid_date(r.get("sale_date"))
    ltcg = _holding_months(buy, sold) >= 24 or (buy and sold) and r.get("force_ltcg")
    buy_year = int(buy[:4]) if buy else 2025
    out = {"description": str(r.get("description") or f"Property {idx}"),
           "buy_date": buy, "sale_date": sold, "sale": sale,
           "cost": cost, "improvement": improv, "expenses": exp,
           "quarter": _quarter(sold)}
    if not ltcg:
        out.update(type="stcg_apprate", gain=sale - cost - improv - exp,
                   note=f"Held {_holding_months(buy, sold):.1f} months (<24) → STCG at slab rate")
        return out
    gain_wo = sale - cost - improv - exp                       # 12.5% without indexation
    if buy and sold < RATE_CHANGE_DATE:
        out.update(type="ltcg_indexed20", gain=gain_wo, note="Sold before 23-07-2024 → 20% with indexation")
        idx_cost = _indexed_cost(cost, buy_year) + _indexed_cost(improv, buy_year)
        out["gain"] = sale - idx_cost - exp
        out["indexed_cost"] = idx_cost
        return out
    # sold on/after 23-07-2024: option available only if acquired before that date
    if buy and buy < RATE_CHANGE_DATE:
        idx_cost = _indexed_cost(cost, buy_year) + _indexed_cost(improv, buy_year)
        gain_idx = sale - idx_cost - exp
        choice = str(r.get("indexation_option", "auto"))
        tax_wo = max(0, gain_wo) * RATE_LTCG_OTHER
        tax_idx = max(0, gain_idx) * RATE_LTCG_INDEXED
        use_idx = (tax_idx <= tax_wo) if choice == "auto" else choice == "with"
        out["gain_wo_index"], out["indexed_cost"] = gain_wo, idx_cost
        out["gain_with_index"] = gain_idx
        out["note"] = (f"12.5% w/o indexation (tax Rs. {tax_wo:,.0f}) vs 20% with indexation "
                       f"(tax Rs. {tax_idx:,.0f}) → {'20% indexed' if use_idx else '12.5%'} chosen")
        if use_idx:
            out.update(type="ltcg_indexed20", gain=gain_idx)
        else:
            out.update(type="ltcg_125", gain=gain_wo)
        return out
    out.update(type="ltcg_125", gain=gain_wo,
               note="Acquired on/after 23-07-2024 → 12.5% without indexation")
    return out


def _other_row(r: dict, idx: int, long: bool) -> dict:
    label = LBL_LONG if long else LBL_SHORT
    with field_scope(f"{label} CG row {idx} ({r.get('description') or 'asset'})"):
        sale = _g(r, "sale_consideration")
        cost = _g(r, "cost_of_acquisition")
        improv = _g(r, "cost_of_improvement")
        exp = _g(r, "expenses")
    buy = _valid_date(r.get("buy_date"))
    sold = _valid_date(r.get("sale_date"))
    base = {"description": str(r.get("description") or f"Asset {idx}"),
            "buy_date": buy, "sale_date": sold, "quarter": _quarter(sold), "sale": sale,
            "cost": cost, "improvement": improv, "expenses": exp}
    if not long:
        base.update(type="stcg_apprate", gain=sale - cost - improv - exp)
        return base
    buy_year = int(buy[:4]) if buy else 2025
    if sold and sold < RATE_CHANGE_DATE:
        idx_cost = _indexed_cost(cost, buy_year) + _indexed_cost(improv, buy_year)
        base.update(type="ltcg_indexed20", gain=sale - idx_cost - exp,
                    indexed_cost=idx_cost, note="Sold before 23-07-2024 → 20% with indexation")
        return base
    base.update(type="ltcg_125", gain=sale - cost - improv - exp,
                note="12.5% without indexation (sold on/after 23-07-2024)")
    return base


LBL_LONG, LBL_SHORT = "Other LTCG", "Other STCG"


# ---------------------------------------------------------------------------
# Exemptions & loss set-off
# ---------------------------------------------------------------------------

EXEMPTION_SECTIONS = {
    "54": "Investment in residential house (against LTCG on house)",
    "54B": "Investment in agricultural land",
    "54EC": "NHAI/REC bonds (against LTCG, ≤ Rs. 50 lakh)",
    "54EE": "Notified fund units (against any LTCG, ≤ Rs. 50 lakh)",
    "54F": "Investment in residential house (against LTCG on other assets)",
}


def compute_capital_gains(data: dict) -> dict:
    cg_in = data.get("capital_gains") or {}
    notes = []

    eq_lt = [_equity_lt_row(r, i + 1) for i, r in enumerate(cg_in.get("equity_lt") or [])]
    eq_st = [_equity_st_row(r, i + 1) for i, r in enumerate(cg_in.get("equity_st") or [])]
    land = [_land_row(r, i + 1) for i, r in enumerate(cg_in.get("land_building") or [])]
    o_st = [_other_row(r, i + 1, False) for i, r in enumerate(cg_in.get("other_st") or [])]
    o_lt = [_other_row(r, i + 1, True) for i, r in enumerate(cg_in.get("other_lt") or [])]
    all_rows = eq_lt + eq_st + land + o_st + o_lt

    # class totals before exemptions
    def tot(pred):
        return sum(r["gain"] for r in all_rows if pred(r))

    # ---- intra-head loss set-off (aggregate model, transparent for the user) ----
    stcg_111a = tot(lambda r: r in eq_st)
    ltcg_112a_gross = tot(lambda r: r in eq_lt)
    stcg_apprate = tot(lambda r: r.get("type") == "stcg_apprate")
    ltcg_125 = tot(lambda r: r.get("type") == "ltcg_125")
    ltcg_idx20 = tot(lambda r: r.get("type") == "ltcg_indexed20")
    ltcg_other_gross = ltcg_125 + ltcg_idx20

    setoff_notes = []
    neg_st = sum(g for g in (r["gain"] for r in eq_st) if g < 0) + \
        sum(g for g in (r["gain"] for r in all_rows if r.get("type") == "stcg_apprate") if g < 0)
    pos_st_111a = sum(g for g in (r["gain"] for r in eq_st) if g > 0)
    pos_st_apprate = sum(g for g in (r["gain"] for r in all_rows if r.get("type") == "stcg_apprate") if g > 0)
    neg_lt = sum(g for g in (r["gain"] for r in eq_lt + land + o_lt) if g < 0)
    pos_lt_112a = sum(g for g in (r["gain"] for r in eq_lt) if g > 0)
    pos_lt_125 = sum(r["gain"] for r in land + o_lt
                     if r.get("type") == "ltcg_125" and r["gain"] > 0)
    pos_lt_idx = sum(r["gain"] for r in land + o_lt
                     if r.get("type") == "ltcg_indexed20" and r["gain"] > 0)

    stcg_loss_left = -neg_st
    use = min(stcg_loss_left, pos_st_111a); pos_st_111a -= use; stcg_loss_left -= use
    if use: setoff_notes.append(f"ST loss Rs. {use:,} set off against STCG u/s 111A.")
    use = min(stcg_loss_left, pos_st_apprate); pos_st_apprate -= use; stcg_loss_left -= use
    if use: setoff_notes.append(f"ST loss Rs. {use:,} set off against STCG at slab rate.")
    # remaining ST loss can set off LT gains (ST loss → any gain)
    if stcg_loss_left > 0:
        use = min(stcg_loss_left, pos_lt_112a); pos_lt_112a -= use; stcg_loss_left -= use
        if use: setoff_notes.append(f"ST loss Rs. {use:,} set off against LTCG u/s 112A.")
        use = min(stcg_loss_left, pos_lt_125); pos_lt_125 -= use; stcg_loss_left -= use
        if use: setoff_notes.append(f"ST loss Rs. {use:,} set off against LTCG (12.5%).")
        use = min(stcg_loss_left, pos_lt_idx); pos_lt_idx -= use; stcg_loss_left -= use
        if use: setoff_notes.append(f"ST loss Rs. {use:,} set off against LTCG (20% indexed).")
    ltcg_loss_left = -neg_lt
    if ltcg_loss_left > 0:                     # LT loss → LT gains only
        use = min(ltcg_loss_left, pos_lt_112a); pos_lt_112a -= use; ltcg_loss_left -= use
        if use: setoff_notes.append(f"LT loss Rs. {use:,} set off against LTCG u/s 112A.")
        use = min(ltcg_loss_left, pos_lt_125); pos_lt_125 -= use; ltcg_loss_left -= use
        if use: setoff_notes.append(f"LT loss Rs. {use:,} set off against LTCG (12.5%).")
        use = min(ltcg_loss_left, pos_lt_idx); pos_lt_idx -= use; ltcg_loss_left -= use
        if use: setoff_notes.append(f"LT loss Rs. {use:,} set off against LTCG (20% indexed).")

    # ---- exemptions ----
    exemptions = []
    for ex in cg_in.get("exemptions") or []:
        sec = str(ex.get("section", ""))
        with field_scope(f"CG exemption u/s {sec}"):
            amt = _g(ex, "amount")
        if amt <= 0:
            continue
        if sec == "54EC":
            if amt > CAP_54EC:
                setoff_notes.append(f"54EC capped at Rs. {CAP_54EC:,}.")
                amt = CAP_54EC
            use = min(amt, pos_lt_idx); pos_lt_idx -= use
            rest = amt - use
            use2 = min(rest, pos_lt_125); pos_lt_125 -= use2; rest -= use2
            use3 = min(rest, pos_lt_112a); pos_lt_112a -= use3
            applied = use + use2 + use3
        elif sec == "54EE":
            if amt > CAP_54EE:
                setoff_notes.append(f"54EE capped at Rs. {CAP_54EE:,}.")
                amt = CAP_54EE
            use = min(amt, pos_lt_idx); pos_lt_idx -= use
            rest = amt - use
            use2 = min(rest, pos_lt_125); pos_lt_125 -= use2; rest -= use2
            use3 = min(rest, pos_lt_112a); pos_lt_112a -= use3
            applied = use + use2 + use3
        elif sec == "54":
            applied = min(amt, pos_lt_idx + pos_lt_125)
            use = min(amt, pos_lt_idx); pos_lt_idx -= use
            use2 = min(amt - use, pos_lt_125); pos_lt_125 -= use2
        elif sec == "54F":
            applied = min(amt, pos_lt_112a)
            pos_lt_112a -= applied
        else:
            applied = 0
            setoff_notes.append(f"Exemption u/s {sec} not auto-applied — check manually.")
        if applied > 0:
            exemptions.append({"section": sec, "claimed": amt, "applied": applied,
                               "label": EXEMPTION_SECTIONS.get(sec, sec)})
    total_exempt_applied = sum(e["applied"] for e in exemptions)

    # ---- final heads ----
    ltcg_112a_after = pos_lt_112a
    taxable_112a_gain = max(0, ltcg_112a_after - EXEMPT_112A)
    exempt_112a_used = min(ltcg_112a_after, EXEMPT_112A)
    stcg_111a = pos_st_111a
    stcg_apprate = pos_st_apprate
    ltcg_125 = pos_lt_125
    ltcg_idx20 = pos_lt_idx

    # ---- quarterly accrual (using post-set-off class level is too complex for
    # row mapping; quarter split is computed on positive row gains pro-rata) ----
    def quarterly(rows2):
        qmap = {q: 0 for q, _, _ in QUARTERS}
        pos_total = sum(max(0, r["gain"]) for r in rows2)
        net_total = sum(r["gain"] for r in rows2)
        if pos_total <= 0:
            return qmap
        ratio = net_total / pos_total
        for r in rows2:
            qmap[r["quarter"]] += max(0, r["gain"]) * ratio
        return {k: round(v) for k, v in qmap.items()}

    res = {
        "rows": {"equity_lt": eq_lt, "equity_st": eq_st, "land_building": land,
                 "other_st": o_st, "other_lt": o_lt},
        "setoff_notes": setoff_notes + notes,
        "exemptions_applied": exemptions,
        "exemptions_total": total_exempt_applied,
        "stcg_111a": round(stcg_111a),
        "stcg_apprate": round(stcg_apprate),
        "ltcg_112a_gross": round(ltcg_112a_after),
        "ltcg_112a_taxable": round(taxable_112a_gain),
        "exempt_112a_used": round(exempt_112a_used),
        "ltcg_other_125": round(ltcg_125),
        "ltcg_indexed_20": round(ltcg_idx20),
        "ltcg_other_taxable": round(ltcg_125 + ltcg_idx20),
        "quarterly": {
            "111a": quarterly(eq_st),
            "112a": quarterly(eq_lt),
            "apprate_st": quarterly([r for r in all_rows if r.get("type") == "stcg_apprate"]),
            "ltcg_rate": quarterly(land + o_lt),
        },
        "cg_loss_carry_st": round(max(0, stcg_loss_left)),
        "cg_loss_carry_lt": round(max(0, ltcg_loss_left)),
        "worksheet": [],
    }
    res["total_cg"] = (res["stcg_111a"] + res["stcg_apprate"] + res["ltcg_112a_taxable"] +
                       res["ltcg_other_taxable"])
    res["total_cg_incl_exempt112a"] = res["total_cg"] + res["exempt_112a_used"]

    # ITR-1 eligibility: only 112A within the exemption limit, nothing else
    other_cg = (res["stcg_111a"] or res["stcg_apprate"] or res["ltcg_other_taxable"] or
                res["ltcg_112a_taxable"] or res["cg_loss_carry_lt"] or res["cg_loss_carry_st"])
    res["itr1_ok"] = not other_cg and res["ltcg_112a_gross"] >= 0
    res["has_any_cg"] = bool(all_rows)

    # worksheet rows for the computation report
    def w(txt, amt):
        res["worksheet"].append({"label": txt, "amount": round(amt)})
    if eq_st:
        w("STCG u/s 111A (equity/equity-MF, STT paid) @ 20%", stcg_111a)
    if stcg_apprate:
        w("STCG at slab rates (other assets)", stcg_apprate)
    if ltcg_112a_after or eq_lt:
        w(f"LTCG u/s 112A @ 12.5% (gross Rs. {ltcg_112a_after:,} − exemption Rs. {exempt_112a_used:,})",
          taxable_112a_gain)
    if ltcg_125:
        w("LTCG @ 12.5% without indexation", ltcg_125)
    if ltcg_idx20:
        w("LTCG @ 20% with indexation", ltcg_idx20)
    for e in exemptions:
        w(f"Less: exemption u/s {e['section']} — {e['label']}", -e["applied"])
    return res


# ---------------------------------------------------------------------------
# Excel / CSV import-export of the scrip-wise schedules
# ---------------------------------------------------------------------------

XLSX_SHEETS = {
    "equity_st": ("Equity STCG (111A)",
                  ["scrip", "isin", "buy_date", "sale_date", "qty", "cost_of_acquisition",
                   "sale_consideration", "expenses"]),
    "equity_lt": ("Equity LTCG (112A)",
                  ["scrip", "isin", "buy_date", "sale_date", "qty", "cost_of_acquisition",
                   "fmv_31_01_2018", "sale_consideration", "expenses"]),
    "land_building": ("Land & Building",
                      ["description", "buy_date", "sale_date", "cost_of_acquisition",
                       "cost_of_improvement", "sale_consideration", "expenses",
                       "indexation_option"]),
    "other_st": ("Other STCG", ["description", "buy_date", "sale_date",
                                "cost_of_acquisition", "sale_consideration", "expenses"]),
    "other_lt": ("Other LTCG", ["description", "buy_date", "sale_date",
                                "cost_of_acquisition", "cost_of_improvement",
                                "sale_consideration", "expenses"]),
    "exemptions": ("Exemptions", ["section", "amount", "date"]),
}


def export_xlsx(data: dict) -> bytes:
    """Write the capital_gains block to an .xlsx workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    cg_in = data.get("capital_gains") or {}
    wb = Workbook()
    first = True
    for key, (title, cols) in XLSX_SHEETS.items():
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = title[:31]
        ws.append([title])
        ws["A1"].font = Font(bold=True, size=13)
        ws.append(["# " + c for c in cols])
        for cell in ws[2]:
            cell.font = Font(bold=True)
        for r in cg_in.get(key) or []:
            ws.append([r.get(c, "") for c in cols])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def import_xlsx(payload: bytes) -> dict:
    """Read the workbook back into the capital_gains model. Also accepts the
    department's CSV_112A / CSV_115AD layout via :func:`import_csv_112a`."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    out = {}
    for key, (title, cols) in XLSX_SHEETS.items():
        ws = None
        for name in wb.sheetnames:
            if name.lower().startswith(title.split("(")[0].strip().lower()[:12]):
                ws = wb[name]
                break
        if ws is None:
            continue
        rows = []
        grid = list(ws.iter_rows(values_only=True))
        header_idx = None
        for i, row in enumerate(grid[:5]):
            if row and any(str(c or "").strip().startswith("#") for c in row):
                header_idx = i
                break
        if header_idx is None:
            continue
        for row in grid[header_idx + 1:]:
            if not row or all(v in (None, "") for v in row):
                continue
            item = {}
            for c, v in zip(cols, row):
                if v is None:
                    continue
                if c in ("buy_date", "sale_date", "date") and hasattr(v, "strftime"):
                    v = v.strftime("%Y-%m-%d")
                item[c] = v
            rows.append(item)
        if rows:
            out[key] = rows
    return {"capital_gains": out}


def import_csv_112a(text: str) -> dict:
    """Parse the department's CSV_112A/CSV_115AD schedule into equity_lt rows."""
    rows = []
    reader = csv.reader(text.splitlines())
    header = None
    for record in reader:
        if not record or not any(record):
            continue
        if header is None:
            header = [c.lower() for c in record]
            continue
        joined = record + [""] * (12 - len(record))
        rows.append({
            "scrip": joined[2].strip(), "isin": joined[1].strip(),
            "qty": joined[3].strip(), "cost_of_acquisition": joined[6].replace(",", "").strip(),
            "fmv_31_01_2018": joined[7].replace(",", "").strip(),
            "sale_consideration": (joined[4].replace(",", "").strip()),
            "expenses": "0", "buy_date": "", "sale_date": "",
        })
    return {"capital_gains": {"equity_lt": rows}}


# ---------------------------------------------------------------------------
# JSON <-> text helpers used by the web/API layer
# ---------------------------------------------------------------------------

def cg_template_json() -> str:
    """A ready-to-fill JSON template users can also paste/import."""
    return json.dumps({
        "capital_gains": {
            "equity_st": [{"scrip": "INFY", "isin": "INE009A01021", "buy_date": "2025-05-10",
                           "sale_date": "2025-08-20", "cost_of_acquisition": 100000,
                           "sale_consideration": 130000, "expenses": 0}],
            "equity_lt": [{"scrip": "SBI Mutual Fund", "isin": "", "buy_date": "2019-01-15",
                           "sale_date": "2026-01-10", "cost_of_acquisition": 200000,
                           "fmv_31_01_2018": 0, "sale_consideration": 420000, "expenses": 0}],
            "land_building": [{"description": "Plot at Surat", "buy_date": "2010-06-01",
                               "sale_date": "2025-11-15", "cost_of_acquisition": 1500000,
                               "cost_of_improvement": 0, "sale_consideration": 6000000,
                               "expenses": 50000, "indexation_option": "auto"}],
            "other_st": [], "other_lt": [], "exemptions": [],
        }}, indent=2)
